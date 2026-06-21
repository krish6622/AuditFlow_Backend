"""One-time migration of the legacy customer registers into the customers table.

Elangovan Associates kept two spreadsheets — a GST client database and an
Income-Tax client database. This module reads both, maps the columns onto the
``customers`` master, and upserts them with the duplicate-detection rules below.
It is invoked once during development via ``python -m app.cli seed-customers``;
it is NOT exposed to end users (no import/upload UI exists by design).

Sources (first that resolves wins per register):
  1. A local file (CSV or XLSX) under ``CUSTOMER_SEED_DIR`` — air-gap friendly.
     The file is matched by name keyword (``gst`` / ``income``/``it client``),
     so the original "GST CLIENT DATABASE.xlsx" / "IT CLIENT DATABASE.xlsx" work.
  2. The Google Sheet CSV export URL (requires the sheet shared as link-viewable).

The reader is tolerant of real-world sheets: it scans for the true header row
(skipping title banners), reads every worksheet, preserves date cells, and
normalises Excel's float-formatted numbers (``9842024145.0`` -> ``9842024145``).
Credential columns (passwords, login IDs, OTP mode, …) have no mapping and are
simply ignored — they are never imported.

Duplicate detection (priority): GST number > PAN number > Mobile > Email.
  - Match found  -> fill only the fields that are currently blank (preserve
                    existing data); if nothing is blank, skip (exact duplicate).
  - No match     -> create a new customer with the next CUS-XXXX code.
Because matching keys are checked on every run, re-running never creates
duplicates — the migration is idempotent.
"""
from __future__ import annotations

import csv
import io
import os
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.features.customers.repository import CustomerRepository
from app.models.customer import Customer
from app.models.enums import CustomerType
from app.models.user import User

logger = get_logger(__name__)

# Spreadsheet IDs from the legacy registers (document identifiers, not secrets).
# Override with CUSTOMER_GST_SHEET_ID / CUSTOMER_IT_SHEET_ID if they ever change.
GST_SHEET_ID = os.getenv(
    "CUSTOMER_GST_SHEET_ID", "1nSXZnzC4gzn2nFbPSFv-SMhJrPMyqpcrscUsF3wHpMw"
)
IT_SHEET_ID = os.getenv(
    "CUSTOMER_IT_SHEET_ID", "1T_ztN3dGrBtz49Q3PO2iSkijeWpLojAzPa7v2RG-mSU"
)
# Where local CSV/XLSX exports are looked for first.
SEED_DIR = Path(os.getenv("CUSTOMER_SEED_DIR", "data/customers"))

_CSV_EXPORT = "https://docs.google.com/spreadsheets/d/{id}/export?format=csv"

# Canonical field -> header aliases (lowercased). Order matters: more specific
# fields (address_line_2, alternate_mobile_number) resolve before their broader
# cousins so a column is never claimed by the wrong field.
_FIELD_ALIASES: list[tuple[str, tuple[str, ...]]] = [
    ("gst_number", ("gstin", "gst no", "gst number", "gst")),
    ("pan_number", ("pan no", "pan number", "pan")),
    ("aadhaar_number", ("aadhaar", "aadhar", "uid")),
    ("alternate_mobile_number", ("alternate", "alt mobile", "alt phone", "secondary", "mobile 2", "phone 2", "contact 2")),
    ("mobile_number", ("mobile no", "mobile", "phone", "contact no", "contact number", "cell", "whatsapp", "ph no")),
    ("email", ("email", "e-mail", "mail id", "gmail", "mail")),
    ("date_of_birth", ("date of birth", "dob", "birth")),
    ("business_name", ("business name", "firm name", "trade name", "company name", "firm", "business", "trade")),
    ("proprietor_name", ("proprietor", "prop name", "owner", "contact person", "authorized person", "partner")),
    ("address_line_2", ("address 2", "address line 2", "address2", "street 2")),
    ("address_line_1", ("address 1", "address line 1", "address", "addr", "door no", "street")),
    ("city", ("city", "town", "place", "district")),
    ("state", ("state",)),
    ("pincode", ("pincode", "pin code", "postal", "zip", "pin")),
    ("remarks", ("remark", "remarks", "notes", "note", "comment", "status detail")),
    # client_name last so "business name"/"contact person" win their columns first.
    ("client_name", ("client name", "customer name", "name of the client", "name of assessee", "assessee", "party name", "client", "name")),
]

# String column limits in the customers model (None = unbounded Text).
_MAX_LEN: dict[str, int] = {
    "client_name": 255, "business_name": 255, "proprietor_name": 255,
    "mobile_number": 40, "alternate_mobile_number": 40, "email": 255,
    "gst_number": 20, "pan_number": 20, "aadhaar_number": 20,
    "city": 120, "state": 120, "pincode": 12,
}

_DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d",
    "%d.%m.%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y",
)


@dataclass
class ImportSummary:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)

    def merge(self, other: "ImportSummary") -> None:
        self.created += other.created
        self.updated += other.updated
        self.skipped += other.skipped
        self.errors.extend(other.errors)

    def __str__(self) -> str:
        return f"created={self.created} updated={self.updated} skipped={self.skipped} errors={len(self.errors)}"


class SheetUnavailableError(RuntimeError):
    """Raised when a register can't be read (sheet not shared, no local file)."""


# --------------------------------------------------------------------------- #
# Cell / header helpers
# --------------------------------------------------------------------------- #
def _clean_cell(value: object) -> object:
    """Normalise a raw cell: keep dates as objects, render integral floats
    without the trailing ``.0`` (Excel stores numbers as floats), trim text."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, bool):
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _score_header_row(cells: tuple) -> int:
    """How many distinct canonical fields the row's cells could be headers for.
    Used to locate the real header row beneath any title banner."""
    norms = [str(c).strip().lower() for c in cells if c not in (None, "")]
    score = 0
    for _field, aliases in _FIELD_ALIASES:
        if any((n == a or a in n) for n in norms for a in aliases):
            score += 1
    return score


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------- #
# Loading rows
# --------------------------------------------------------------------------- #
def _local_file(customer_type: CustomerType) -> Path | None:
    """First local export matching the register, by filename keyword."""
    if not SEED_DIR.exists():
        return None
    candidates = sorted(SEED_DIR.glob("*.xlsx")) + sorted(SEED_DIR.glob("*.csv"))
    for path in candidates:
        name = path.name.lower()
        if customer_type == CustomerType.GST:
            if "gst" in name:
                return path
        else:
            if "gst" not in name and any(
                k in name for k in ("income", "it client", "incometax", "income_tax")
            ):
                return path
    return None


def _read_csv_text(text: str) -> list[dict[str, object]]:
    reader = csv.DictReader(io.StringIO(text))
    return [
        {(k or "").strip(): (v or "").strip() for k, v in row.items()}
        for row in reader
    ]


def _read_xlsx(path: Path) -> list[dict[str, object]]:
    """Read every worksheet, locating the true header row in each (sheets may
    carry a title banner above the header). Returns one dict per data row."""
    try:
        from openpyxl import load_workbook  # optional dependency
    except ImportError as exc:  # pragma: no cover - depends on env
        raise SheetUnavailableError(
            f"Reading {path} needs openpyxl (pip install openpyxl), or export the sheet as CSV."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[dict[str, object]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Locate the header row: the best-scoring of the first 15 rows (>= 2 hits).
        header_idx, best = None, 1
        for i in range(min(15, len(rows))):
            score = _score_header_row(rows[i])
            if score > best:
                best, header_idx = score, i
        if header_idx is None:
            logger.warning("Sheet '%s' in %s has no recognizable header — skipped", ws.title, path.name)
            continue
        headers = [str(_clean_cell(c)) for c in rows[header_idx]]
        for raw in rows[header_idx + 1:]:
            row: dict[str, object] = {}
            for j, h in enumerate(headers):
                if h and j < len(raw):
                    row[h] = _clean_cell(raw[j])
            if any(v not in (None, "") for v in row.values()):
                out.append(row)
    return out


def _fetch_sheet_csv(spreadsheet_id: str) -> list[dict[str, object]]:
    url = _CSV_EXPORT.format(id=spreadsheet_id)
    req = urllib.request.Request(url, headers={"User-Agent": "AuditFlow-seed/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:  # noqa: S310 (trusted Google URL)
        payload = resp.read().decode("utf-8", errors="replace")
    if payload.lstrip().lower().startswith(("<!doctype", "<html")):
        raise SheetUnavailableError(
            f"Sheet {spreadsheet_id} is not publicly readable (got a login page). "
            "Share it as 'Anyone with the link -> Viewer', or drop a CSV/XLSX export "
            f"into {SEED_DIR}/."
        )
    return _read_csv_text(payload)


def _load_rows(customer_type: CustomerType, spreadsheet_id: str) -> list[dict[str, object]]:
    local = _local_file(customer_type)
    if local is not None:
        logger.info("Reading %s register from local file %s", customer_type.value, local.name)
        return _read_xlsx(local) if local.suffix.lower() == ".xlsx" else _read_csv_text(
            local.read_text(encoding="utf-8-sig")
        )
    logger.info("Fetching %s register from Google Sheet %s", customer_type.value, spreadsheet_id)
    return _fetch_sheet_csv(spreadsheet_id)


# --------------------------------------------------------------------------- #
# Mapping
# --------------------------------------------------------------------------- #
def _build_header_map(headers: list[str]) -> dict[str, str]:
    """Map canonical field -> source header, resolving each field to the first
    unclaimed column whose name matches one of its aliases."""
    normalized = {h: h.strip().lower() for h in headers if h and h.strip()}
    claimed: set[str] = set()
    mapping: dict[str, str] = {}
    for field_name, aliases in _FIELD_ALIASES:
        match = None
        for want_exact in (True, False):  # exact alias first, then substring
            for header, norm in normalized.items():
                if header in claimed:
                    continue
                if any((norm == a) if want_exact else (a in norm) for a in aliases):
                    match = header
                    break
            if match:
                break
        if match:
            mapping[field_name] = match
            claimed.add(match)
    return mapping


def _row_to_fields(row: dict[str, object], header_map: dict[str, str]) -> dict[str, object]:
    fields: dict[str, object] = {}
    for field_name, header in header_map.items():
        raw = row.get(header)
        if raw is None or raw == "":
            continue
        if field_name == "date_of_birth":
            parsed = _coerce_date(raw)
            if parsed is not None:
                fields[field_name] = parsed
            continue
        text = (raw if isinstance(raw, str) else str(raw)).strip()
        if not text:
            continue
        if field_name == "email":
            text = text.lower()
        cap = _MAX_LEN.get(field_name)
        if cap is not None and len(text) > cap:
            text = text[:cap]
        fields[field_name] = text
    return fields


def _ordered_headers(rows: list[dict[str, object]]) -> list[str]:
    seen: list[str] = []
    for row in rows:
        for key in row:
            if key not in seen:
                seen.append(key)
    return seen


# --------------------------------------------------------------------------- #
# Upsert with duplicate detection
# --------------------------------------------------------------------------- #
def _import_register(
    db: Session,
    repo: CustomerRepository,
    *,
    organization_id,
    created_by,
    customer_type: CustomerType,
    rows: list[dict[str, object]],
) -> ImportSummary:
    summary = ImportSummary()
    if not rows:
        return summary
    header_map = _build_header_map(_ordered_headers(rows))
    if "client_name" not in header_map:
        summary.errors.append(
            f"{customer_type.value}: no client-name column found in headers {_ordered_headers(rows)[:20]}"
        )
        return summary
    logger.info("%s register header mapping: %s", customer_type.value, header_map)

    for idx, row in enumerate(rows, start=2):
        fields = _row_to_fields(row, header_map)
        if not fields.get("client_name"):
            continue  # blank / non-client line

        try:
            # Each row in its own savepoint so one bad row can't poison the batch.
            with db.begin_nested():
                existing = repo.find_duplicate(
                    organization_id=organization_id,
                    gst_number=fields.get("gst_number"),  # type: ignore[arg-type]
                    pan_number=fields.get("pan_number"),  # type: ignore[arg-type]
                    mobile_number=fields.get("mobile_number"),  # type: ignore[arg-type]
                    email=fields.get("email"),  # type: ignore[arg-type]
                )
                if existing is None:
                    seq = repo.next_code_sequence(organization_id=organization_id)
                    customer = Customer(
                        organization_id=organization_id,
                        customer_code=f"CUS-{seq:04d}",
                        customer_type=customer_type,
                        created_by=created_by,
                        is_active=True,
                        **fields,
                    )
                    repo.add(customer)
                    db.flush()
                    summary.created += 1
                else:
                    changed = False
                    for field_name, value in fields.items():
                        if field_name == "client_name":
                            continue  # never rename an existing client on import
                        if getattr(existing, field_name) in (None, ""):
                            setattr(existing, field_name, value)
                            changed = True
                    if changed:
                        db.flush()
                        summary.updated += 1
                    else:
                        summary.skipped += 1
        except Exception as exc:  # noqa: BLE001 - capture and continue the batch
            name = fields.get("client_name")
            summary.errors.append(f"{customer_type.value} row {idx} ({name}): {exc}")
            logger.warning("Skipped %s row %d (%s): %s", customer_type.value, idx, name, exc)
    return summary


def import_customers(db: Session, admin: User) -> ImportSummary:
    """Import both registers into ``admin``'s organization. Idempotent."""
    if admin.organization_id is None:
        raise ValueError("Seed admin has no organization")
    repo = CustomerRepository(db)
    total = ImportSummary()
    for customer_type, sheet_id in (
        (CustomerType.GST, GST_SHEET_ID),
        (CustomerType.INCOME_TAX, IT_SHEET_ID),
    ):
        try:
            rows = _load_rows(customer_type, sheet_id)
        except SheetUnavailableError as exc:
            total.errors.append(str(exc))
            logger.warning("Skipping %s register: %s", customer_type.value, exc)
            continue
        logger.info("%s register: %d data rows read", customer_type.value, len(rows))
        result = _import_register(
            db,
            repo,
            organization_id=admin.organization_id,
            created_by=admin.id,
            customer_type=customer_type,
            rows=rows,
        )
        logger.info("%s register import: %s", customer_type.value, result)
        total.merge(result)
    db.commit()
    return total
